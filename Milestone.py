import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from datetime import date
from milestone.veridia import *
from milestone.EwsLig import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from milestone.Eligo import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re  # Added for date extraction

COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"

cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write the DataFrame to Excel, starting from row 1 to leave space for the title
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=1)
        
        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Define a format for the title row with yellow background
        title_format = workbook.add_format({
            'bold': True,
            'bg_color': 'yellow',
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Merge cells across the first row for the title
        worksheet.merge_range(0, 0, 0, len(df.columns)-1, f'Overall Project Report ({date.today()})', title_format)
        
    return output.getvalue()

# Function to extract date from filename
def extract_date(filename):
    match = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
    if match:
        day, month, year = map(int, match.groups())
        return datetime(year, month, day)
    return None

# Create Excel file content

foundewslig = False
foundeiden = False
foundeligo = False
foundverdia = False
foundeligog = False
foundeligoh = False
foundveridiaf4 = False
foundveridiaf5 = False

veridia = None
ews_lig = None
eligo = None

def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files"]
    
files = get_cos_files()
st.write(files)

today = date.today()
prev_month = today - relativedelta(months=1)

foundverdia = False

month_year = today.strftime("%m-%Y")
prev_month_year = prev_month.strftime("%m-%Y")

# Filter files for those on or after the 10th of the current month
current_year = today.year
current_month = today.month
cutoff_day = 10

files_after_or_on_10th = []
for f in files:
    file_date = extract_date(f)
    if file_date and file_date.year == current_year and file_date.month == current_month and file_date.day >= cutoff_day:
        files_after_or_on_10th.append(f)

#=============VERIDIA================
for file in files_after_or_on_10th:
    try:
        if file.startswith("Veridia") and "Structure Work Tracker" in file and month_year in file:
            st.write("✅ Current month:", file)
            response = cos_client.get_object(Bucket="projectreportnew", Key=file)
            veridia = ProcessMilestone1(io.BytesIO(response['Body'].read()))
            foundverdia = True
            break
    except Exception as e:
        st.info(e)

if not foundverdia:
    for file in files:
        try:
            if file.startswith("Veridia") and "Structure Work Tracker" in file and prev_month_year in file:
                st.write("🕓 Previous month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                veridia = ProcessMilestone1(io.BytesIO(response['Body'].read()))
                st.write(veridia)
                break
        except Exception as e:
            st.error(e)

#=============VERIDIA================

#===========EWS LIG=================
for file in files_after_or_on_10th:
    try:
        if file.startswith("EWS LIG") and "Structure Work Tracker" in file and month_year in file:
            st.write("✅ Current month:", file)
            response = cos_client.get_object(Bucket="projectreportnew", Key=file)
            ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
            foundewslig = True
            break
    except Exception as e:
        st.error(e)

if not foundewslig:
    for file in files:
        try:
            if file.startswith("EWS LIG") and "Structure Work Tracker" in file and prev_month_year in file:
                st.write("🕓 Previous month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
                break
        except Exception as e:
            st.error(e)

#===========EWS LIG=================

#===========ELIGO============
for file in files_after_or_on_10th:
    try:
        if file.startswith("Eligo") and "Structure Work Tracker" in file and month_year in file:
            st.write("✅ Current month:", file)
            response = cos_client.get_object(Bucket="projectreportnew", Key=file)
            eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
            foundeligo = True
            break
    except Exception as e:
        st.info(e)

if not foundeligo:
    for file in files:
        try:
            if file.startswith("Eligo") and "Structure Work Tracker" in file and prev_month_year in file:
                st.write("🕓 Previous month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                break
        except Exception as e:
            st.error(e)
#===========ELIGO============

combined_json = []
missing_sources = []

if veridia is None:
    missing_sources.append("Veridia")
else:
    combined_json += veridia

if ews_lig is None:
    missing_sources.append("EWS_LIG")
else:
    combined_json += ews_lig

if eligo is None:
    missing_sources.append("Eligo")
else:
    combined_json += eligo

if missing_sources:
    st.warning(f"Missing Tow Files from: {', '.join(missing_sources)}")
else:
    st.success("All JSON sources were loaded successfully.")

df = pd.DataFrame(combined_json)
st.write(df)
df['date'] = pd.to_datetime(df['date'])

df['year'] = df['date'].dt.year
df['month'] = df['date'].dt.month

month_map = {
    1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR',
    5: 'MAY', 6: 'JUN', 7: 'JUL', 8: 'AUG',
    9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DEC'
}

df['month_name'] = df['month'].map(month_map)

reverse_month_map = {v: k for k, v in month_map.items()}

selected_year = st.sidebar.selectbox("Select Year", sorted(df['year'].unique()))

selected_month_names = st.sidebar.multiselect(
    "Select Month(s)", 
    options=list(month_map.values()), 
    default=list(month_map.values()) 
)

selected_months = [reverse_month_map[m] for m in selected_month_names]

filtered_df = df[(df['year'] == selected_year) & (df['month'].isin(selected_months))]

st.write(filtered_df)

json_data = filtered_df.to_json(orient='records')
st.write(json_data)

def process_json_data(json_data):
    # Convert data to DataFrame
    df = pd.DataFrame(json_data)
    st.write(df)
    
    # Handle different date formats
    if df['date'].dtype == 'object' and isinstance(df['date'].iloc[0], str):
        df['date_clean'] = df['date'].str.extract(r"Timestamp\('([^']+)'\)")
        df['date_clean'] = pd.to_datetime(df['date_clean'])
    else:
        df['date_clean'] = pd.to_datetime(df['date'])
    
    df['module'] = df['Tower']
    st.write(df['date_clean'])
    unique_months = sorted(df['date_clean'].dt.month.unique())
    st.write(unique_months)
    month_names = []
    month_mapping = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
        7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'
    }
    
    for month_num in unique_months:
        month_names.append(month_mapping[month_num])
    
    result_data = []
    modules = df['module'].unique()
    
    for module in sorted(modules):
        module_data = df[df['module'] == module]
        row = {'Modules': module}
        
        for i, month_num in enumerate(unique_months):
            month_name = month_names[i]
            month_entries = module_data[module_data['date_clean'].dt.month == month_num]
            
            if len(month_entries) == 0:
                row[month_name] = f'No work plan for {month_name.lower()[:3]}'
            else:
                floors = month_entries['floor'].unique()
                row[month_name] = ', '.join(sorted(floors))
        
        result_data.append(row)
    
    return pd.DataFrame(result_data), month_names

def create_excel_file(df, month_names):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    date_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    date_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    today_date = datetime.now().strftime("%B %d, %Y")
    date_cell = ws.cell(row=1, column=1, value=f"Downloaded on: {today_date}")
    date_cell.fill = date_fill
    date_cell.font = date_font
    date_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    headers = ['Modules'] + month_names
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    total_row = len(df) + 3
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    
    for col in range(2, len(headers) + 1):
        month_data = [ws.cell(row=r, column=col).value for r in range(3, total_row)]
        count = sum(1 for item in month_data if item and not item.startswith('No work plan'))
        ws.cell(row=total_row, column=col, value=f"{count} Slabs").font = Font(bold=True)
    
    for col_num in range(1, len(headers) + 1):
        column_letter = ws.cell(row=2, column=col_num).column_letter
        max_length = 0
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

processed_df, month_names = process_json_data(filtered_df)

# After processing all files and creating processed_df
st.subheader("Preview of processed data:")
st.dataframe(processed_df)

# Show detected months
st.info(f"Detected months in your data: {', '.join(month_names)}")

# Generate Excel file only if processed_df is not empty
if not processed_df.empty:
    excel_data = to_excel(processed_df)
    excel_buffer = io.BytesIO(excel_data)
    excel_buffer.seek(0)

    # Download button
    st.download_button(
        label="📥 Download Excel File",
        data=excel_buffer.getvalue(),
        file_name=f"construction_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("No data available to generate Excel file. Please ensure relevant files are uploaded to the COS bucket.")
